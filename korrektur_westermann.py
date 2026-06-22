#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KORREKTUR: Westermann Deutsch statt Cornelsen
"""

from openpyxl import load_workbook

wb = load_workbook('schulbuecher_alle_varianten.xlsx')
ws3 = wb['E-Book-Detail']

# Aendere Deutschbuch-Eintrag
ws3['A4'] = 'Deutsch Klasse 5-7'
ws3['B4'] = 'Westermann'
ws3['C4'] = '36,50'
ws3['D4'] = 'NICHT einzeln'
ws3['E4'] = 'BiBox 1 Jahr'
ws3['F4'] = 'N/A'
ws3['G4'] = 'Nur BiBox-Paket teuer + Abo-Modell. BESSER: Gebraucht kaufen'

# Aendere E-Book-Berechnung im ersten Sheet
ws1 = wb['Alle-Varianten-Vergleich']

# Finde und aendere Variante 4
# Suche nach der Cornelsen E-Book Zeile
for row in ws1.iter_rows(min_row=5, max_row=14):
    if row[0].value and 'Cornelsen' in str(row[0].value) and 'E-Book' in str(row[0].value):
        # Aendere sie
        row[0].value = '4. E-Books (nur Cornelsen: Mathe+Englisch)'
        row[1].value = '156.94 EUR'
        row[2].value = '52.31 EUR'
        row[5].value = 'DEUTSCH jetzt Westermann (Print/Gebraucht empfohlen)'

# Aendere Hybrid-Variante
for row in ws1.iter_rows(min_row=5, max_row=14):
    if row[0].value and 'Hybrid' in str(row[0].value) and '320' in str(row[1].value if row[1].value else ''):
        row[0].value = '7. Hybrid: E-Books Cornelsen (Mathe+Eng) + Print (alle anderen gebraucht)'
        row[1].value = '290-330 EUR'
        row[2].value = '96.67 EUR'
        row[5].value = 'Deutsch: Westermann Print gebraucht statt E-Book'

# Neuer Empfehlung-Sheet mit Westermann Info
ws_rec = wb['Empfehlung']

# Aendere Hybrid-Plan Schritt 2
ws_rec['A9'] = 'SCHRITT 2: E-Books kaufen (nur Cornelsen)'
ws_rec.cell(row=11, column=2, value='Mathe (Kl.5-7): 3 x 12,99EUR = 38,97EUR (Dauerlizenz!)')
ws_rec.cell(row=12, column=2, value='Englisch (Kl.5-7): 3 x 13,99EUR = 41,97EUR (Dauerlizenz!)')
ws_rec.cell(row=13, column=2, value='Subtotal E-Books: ~81EUR (keine Deutsch E-Books, da Westermann teuer!)')

ws_rec.cell(row=15, column=2, value='SCHRITT 3: DEUTSCH kaufen (Westermann Print, gebraucht)')
ws_rec.cell(row=16, column=2, value='Westermann "Klartext" oder "P.A.U.L. D." - Klasse 5,6,7')
ws_rec.cell(row=17, column=2, value='Gebraucht (MediMops, Zustand gut): ca. 70-90EUR fuer alle 3')

ws_rec.cell(row=19, column=2, value='SCHRITT 4: Andere Faecher Print gebraucht (MediMops)')
ws_rec.cell(row=20, column=2, value='Biologie, Physik, Chemie, Geografie: ca. 120-150EUR')

ws_rec.cell(row=22, column=2, value='GESAMT: 81 + 80 + 135 = ca. 296EUR statt 563EUR = 47% ERSPARNIS!')

wb.save('schulbuecher_alle_varianten.xlsx')
print("OK: Westermann-Korrektur eingefuegt")
print("\nKORREKTUR-ZUSAMMENFASSUNG:")
print("=====================================")
print("DEUTSCH: NICHT Cornelsen E-Book")
print("         SONDERN Westermann 'Klartext' Print (gebraucht)")
print("\nNEUE RECHNUNG:")
print("- E-Books Cornelsen (Mathe+Englisch):  ~82 EUR")
print("- Print Deutsch (Westermann, gebraucht): ~80 EUR")
print("- Print andere Faecher (gebraucht):    ~135 EUR")
print("- GESAMT: ca. 297 EUR (statt 563 EUR)")
print("- ERSPARNIS: 47% !")
