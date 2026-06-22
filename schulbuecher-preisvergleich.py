#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Schulbücher-Preisvergleich für Gesamtschule NRW Klassen 5-7
Erstellt eine Excel-Übersicht mit ISBNs und Preisquellen
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ========== SHEET 1: SCHULBÜCHER-ÜBERSICHT ==========
ws = wb.active
ws.title = "Schulbücher"

# Header
ws['A1'] = 'SCHULBÜCHER GESAMTSCHULE NRW KLASSEN 5-7'
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:H1')

# Spalten-Header
headers = ['Klasse', 'Fach', 'Titel', 'Verlag', 'ISBN', 'Typ', 'Ausgabe', 'Hinweis']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Daten - Klasse 5
data_kl5 = [
    ['5', 'Mathematik', 'Dreifach Mathe - Schulbuch', 'Cornelsen', '978-3-06-043746-7', 'Schülerbuch', '2024', 'Differenzierende Ausgabe'],
    ['5', 'Mathematik', 'Dreifach Mathe - Arbeitsheft', 'Cornelsen', '978-3-06-043724-5', 'Arbeitsheft', '2024', 'mit Lösungen'],
    ['5', 'Deutsch', 'Deutschbuch - Schulbuch', 'Cornelsen', '978-3-06-062085-0', 'Schülerbuch', '2020', 'Differenzierende Ausgabe für Gesamtschule'],
    ['5', 'Deutsch', 'Deutschbuch - Arbeitsheft', 'Cornelsen', '---', 'Arbeitsheft', '2020', 'Optional'],
    ['5', 'Englisch', 'English G Lighthouse - Band 1', 'Cornelsen', '978-3-06-032373-9', 'Schülerbuch', '2024', 'Allgemeine Ausgabe, Klasse 5'],
    ['5', 'Englisch', 'English G Lighthouse - Workbook 1', 'Cornelsen', '978-3-06-032505-4', 'Arbeitsheft', '2024', 'mit Audios online'],
    ['5', 'Biologie', 'Natura 1 - Schulbuch', 'Klett', '978-3-12-049200-9', 'Schülerbuch', '2022', 'Allgemeine Ausgabe, Kl. 5/6'],
    ['5', 'Erdkunde', 'Geografie - Schulbuch', 'Cornelsen', '978-3-06-040572-5', 'Schülerbuch', '2023', 'Klasse 5-7'],
    ['5', 'Physik', 'PRISMA Physik 5/6', 'Klett', '978-3-12-068781-8', 'Schülerbuch', '2018', 'Differenzierende Ausgabe NRW'],
]

# Daten - Klasse 6
data_kl6 = [
    ['6', 'Mathematik', 'Dreifach Mathe - Schulbuch', 'Cornelsen', '978-3-06-043747-4', 'Schülerbuch', '2024', 'Differenzierende Ausgabe'],
    ['6', 'Mathematik', 'Dreifach Mathe - Arbeitsheft', 'Cornelsen', '978-3-06-043725-2', 'Arbeitsheft', '2024', 'mit Lösungen'],
    ['6', 'Deutsch', 'Deutschbuch - Schulbuch', 'Cornelsen', '978-3-06-063401-7', 'Schülerbuch', '2020', 'Differenzierende Ausgabe für Gesamtschule'],
    ['6', 'Deutsch', 'Deutschbuch - Arbeitsheft', 'Cornelsen', '---', 'Arbeitsheft', '2020', 'Optional'],
    ['6', 'Englisch', 'English G Lighthouse - Band 2', 'Cornelsen', '978-3-06-032409-5', 'Schülerbuch', '2024', 'Allgemeine Ausgabe, Klasse 6'],
    ['6', 'Englisch', 'English G Lighthouse - Workbook 2', 'Cornelsen', '978-3-06-032641-9', 'Arbeitsheft', '2024', 'mit Audios online'],
    ['6', 'Biologie', 'Natura 1 - Schulbuch', 'Klett', '978-3-12-049200-9', 'Schülerbuch', '2022', 'Allgemeine Ausgabe, Kl. 5/6 (noch!)'],
    ['6', 'Erdkunde', 'Geografie - Schulbuch', 'Cornelsen', '978-3-06-040572-5', 'Schülerbuch', '2023', 'Klasse 5-7'],
    ['6', 'Physik', 'PRISMA Physik 5/6', 'Klett', '978-3-12-068781-8', 'Schülerbuch', '2018', 'Differenzierende Ausgabe NRW'],
]

# Daten - Klasse 7
data_kl7 = [
    ['7', 'Mathematik', 'Dreifach Mathe - Schulbuch', 'Cornelsen', '978-3-06-043748-1', 'Schülerbuch', '2024', 'Differenzierende Ausgabe'],
    ['7', 'Mathematik', 'Dreifach Mathe - Arbeitsheft', 'Cornelsen', '978-3-06-043726-9', 'Arbeitsheft', '2024', 'mit Lösungen'],
    ['7', 'Deutsch', 'Deutschbuch - Schulbuch', 'Cornelsen', '978-3-06-063402-4', 'Schülerbuch', '2020', 'Differenzierende Ausgabe für Gesamtschule'],
    ['7', 'Deutsch', 'Deutschbuch - Arbeitsheft', 'Cornelsen', '---', 'Arbeitsheft', '2020', 'Optional'],
    ['7', 'Englisch', 'English G Lighthouse - Band 3', 'Cornelsen', '978-3-06-032410-1', 'Schülerbuch', '2024', 'Allgemeine Ausgabe, Klasse 7'],
    ['7', 'Englisch', 'English G Lighthouse - Workbook 3', 'Cornelsen', '978-3-06-032642-6', 'Arbeitsheft', '2024', 'mit Audios online'],
    ['7', 'Biologie', 'Natura 2 - Schulbuch', 'Klett', '978-3-12-049207-8', 'Schülerbuch', '2022', 'Allgemeine Ausgabe, Kl. 7-9'],
    ['7', 'Erdkunde', 'Geografie - Schulbuch', 'Cornelsen', '978-3-06-040572-5', 'Schülerbuch', '2023', 'Klasse 5-7'],
    ['7', 'Chemie', 'PRISMA Chemie 7-10', 'Klett', '978-3-12-068509-8', 'Schülerbuch', '2014', 'Differenzierende Ausgabe NRW'],
    ['7', 'Physik', 'PRISMA Physik 7-10', 'Klett', '---', 'Schülerbuch', '2021', 'Differenzierende Ausgabe (recherchieren!)'],
]

row = 4
for data in data_kl5 + data_kl6 + data_kl7:
    for col, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    row += 1

# Spalten-Breite
ws.column_dimensions['A'].width = 7
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 30

# ========== SHEET 2: PREISQUELLEN & ANLEITUNG ==========
ws2 = wb.create_sheet('Preisquellen & Anleitung')

ws2['A1'] = 'PREISVERGLEICH - PORTALE UND TIPPS'
ws2['A1'].font = Font(bold=True, size=14)
ws2.merge_cells('A1:D1')

ws2['A3'] = 'Portale für Preisvergleich:'
ws2['A3'].font = Font(bold=True, size=11)

portale = [
    ('Amazon.de', 'www.amazon.de', 'Große Auswahl, oft guter Preis, Versand ca. 0-3,99€'),
    ('Medimops', 'www.medimops.de', 'Gebrauchte & Neue Bücher, günstiger, Versand ab 3,99€'),
    ('Cornelsen-Verlag', 'www.cornelsen.de', 'Direkt vom Verlag, oftmals mit Rabatten'),
    ('Klett-Verlag', 'www.klett.de', 'Direkt vom Verlag, Original-Ausgaben'),
    ('Thalia', 'www.thalia.de', 'Große Buchhandlung, gute Preise, Versand ab 3,95€'),
    ('Hugendubel', 'www.hugendubel.de', 'Online-Buchhandlung, diverses Sortiment'),
    ('Bücherverkauf.de', 'www.buecherverkauf.de', 'Gebraucht + Neu, günstiger, Versand ab 3€'),
    ('AbeBooks', 'www.abebooks.com', 'Marktplatz (alt + neu), internationale Angebote'),
]

row = 5
for portal, url, hinweis in portale:
    ws2.cell(row=row, column=1, value=portal).font = Font(bold=True)
    ws2.cell(row=row, column=2, value=url)
    ws2.cell(row=row, column=3, value=hinweis)
    ws2.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    ws2.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
    ws2.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
    row += 1

ws2['A15'] = 'Preisvergleich-Strategie:'
ws2['A15'].font = Font(bold=True, size=11)

tipps = [
    '1. ISBNs aus dieser Liste kopieren → in Suchfeld eingeben',
    '2. Neue Bücher (Neuware) bevorzugen für Schulbücher (Vollständigkeit)',
    '3. Versandkosten IMMER einrechnen (wichtig für Gesamtpreis!)',
    '4. Oberhausen: Versand DHL/DPD meist 2-3 Tage',
    '5. MediMops: Gebrauchte Bücher billiger, aber Zustand prüfen',
    '6. Kombi-Angebote: Manchmal billiger mehrere bei einem Portal kaufen',
    '7. Gesamtpreis kalkulieren: Summe aller Bücher + Versand vergleichen',
    '8. Arbeitshefte separat oder im Bundle? Einzeln oft günstiger',
    '9. Kostenlose Zusatzmaterialien: Schulbuchverlage bieten oft PDFs kostenlos',
]

row = 17
for tipp in tipps:
    ws2.cell(row=row, column=1, value=tipp)
    ws2.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    row += 1

ws2['A28'] = 'Kostenlose Zusatzmaterialien:'
ws2['A28'].font = Font(bold=True, size=11)

kostenlos = [
    ('Khan Academy', 'www.khanacademy.org', 'Mathe, Deutsch, Englisch (Video-Lernhilfen)'),
    ('ARD Lernbox', 'www.ardmedienthek.de', 'Kostenlose Videos zu Schulthemen'),
    ('Schlaukopf', 'www.schlaukopf.de', 'Quiz & Übungen für alle Fächer'),
    ('Learnattack', 'www.learnattack.de', 'Kostenlos anmelden, Basis-Zugang'),
    ('YouTube', 'www.youtube.com', 'Kanal: MrWissen2go (Deutsch, Geschichte, Englisch)'),
    ('Cornelsen Lernen-App', 'cornelsen.de/lernen-app', 'Kostenlos, zu ihren Schulbüchern'),
    ('Schulentwicklung NRW', 'schulentwicklung.nrw.de', 'Unterrichtsmaterialien & Lehrpläne kostenlos'),
]

row = 30
for name, url, beschreibung in kostenlos:
    ws2.cell(row=row, column=1, value=name).font = Font(bold=True)
    ws2.cell(row=row, column=2, value=url)
    ws2.cell(row=row, column=3, value=beschreibung)
    ws2.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    ws2.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
    ws2.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
    row += 1

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 45

# ========== SHEET 3: KALKULATIONSVORLAGE ==========
ws3 = wb.create_sheet('Kalkulation')

ws3['A1'] = 'PREISKALKULATION SCHULBÜCHER'
ws3['A1'].font = Font(bold=True, size=14)
ws3.merge_cells('A1:D1')

# Header für Kalkulation
kalk_headers = ['Titel', 'ISBN', 'Menge', 'Amazon', 'MediMops', 'Thalia', 'Günstigster', 'Kosten gesamt']
for col, header in enumerate(kalk_headers, 1):
    cell = ws3.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Beispielzeile als Vorlage
beispiele = [
    ['Dreifach Mathe 5 (Schulbuch)', '978-3-06-043746-7', 1, '', '', '', '=MIN(D4:F4)', '=G4*C4'],
    ['Dreifach Mathe 5 (Arbeitsheft)', '978-3-06-043724-5', 1, '', '', '', '=MIN(D5:F5)', '=G5*C5'],
]

for idx, beispiel in enumerate(beispiele, 4):
    for col, value in enumerate(beispiel, 1):
        ws3.cell(row=idx, column=col, value=value)

# Gesamt-Zeile
ws3['A7'] = 'SUMME (mit Versandkosten):'
ws3['A7'].font = Font(bold=True)
ws3['H7'] = '=SUM(H4:H6)'
ws3['H7'].font = Font(bold=True)

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['E'].width = 12
ws3.column_dimensions['F'].width = 12
ws3.column_dimensions['G'].width = 12
ws3.column_dimensions['H'].width = 15

# Hinweis zur Verwendung
ws3['A10'] = 'Verwendung:'
ws3['A10'].font = Font(bold=True)
ws3['A11'] = '1. Menge anpassen (normalerweise 1 pro Schüler)'
ws3['A12'] = '2. Preise in Amazon/MediMops/Thalia eingeben (von der jeweiligen Website kopieren)'
ws3['A13'] = '3. "Günstigster" zeigt den niedrigsten Preis'
ws3['A14'] = '4. "Kosten gesamt" = Einzelpreis × Menge'
ws3['A15'] = '5. Versandkosten MANUELL unten addieren'
ws3['A16'] = ''
ws3['A17'] = 'Versandkosten (gesamt):'
ws3['A18'] = '€ 0,00'
ws3['A18'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

wb.save('schulbuecher_preisvergleich.xlsx')
print("OK: Excel-Datei erstellt")
