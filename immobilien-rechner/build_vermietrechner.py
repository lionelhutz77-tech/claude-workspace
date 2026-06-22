from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

wb = Workbook()
ws = wb.active
ws.title = "Kalkulation"

# ---- Styles ----
FONT = "Arial"
base = Font(name=FONT, size=10)
bold = Font(name=FONT, size=10, bold=True)
title = Font(name=FONT, size=14, bold=True)
sec = Font(name=FONT, size=11, bold=True, color="FFFFFF")
note_f = Font(name=FONT, size=9, italic=True, color="555555")
white_bold = Font(name=FONT, size=10, bold=True, color="FFFFFF")

input_fill = PatternFill("solid", fgColor="DDEBF7")   # hellblau = Eingabefeld
sec_fill = PatternFill("solid", fgColor="1F4E78")     # dunkelblaue Section-Leiste
res_fill = PatternFill("solid", fgColor="E2EFDA")     # grün = Ergebnis
hdr_fill = PatternFill("solid", fgColor="BDD7EE")

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

EUR = '#,##0 €;(#,##0) €;-'
EUR2 = '#,##0.00 €;(#,##0.00) €;-'
PCT = '0.0%'
PCT2 = '0.00%'
NUM2 = '0.0"x"'

def setc(coord, val, font=base, fill=None, fmt=None, align=None, bd=False, comment=None):
    c = ws[coord]
    c.value = val
    c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if align: c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(align=="wrap"))
    if bd: c.border = border
    if comment: c.comment = Comment(comment, "FlipperImmo")
    return c

def inp(coord, val, fmt=EUR):
    return setc(coord, val, font=Font(name=FONT, size=10, color="00008B"), fill=input_fill, fmt=fmt, align="right", bd=True)

def sectionbar(row, col_start, col_end, text):
    cs = ws[f"{col_start}{row}"]
    cs.value = text
    cs.font = sec
    from openpyxl.utils import column_index_from_string, get_column_letter
    a = column_index_from_string(col_start); b = column_index_from_string(col_end)
    for i in range(a, b+1):
        ws[f"{get_column_letter(i)}{row}"].fill = sec_fill
    ws.merge_cells(f"{col_start}{row}:{col_end}{row}")

# ---- Title ----
setc("B2", "Immobilien-Rechner: Kaufen & Vermieten (Buy & Hold)", title)
setc("B3", "Vergleich monatliche Einnahmen vs. Ausgaben  →  Cashflow & Rendite als Kaufentscheidung", note_f)
setc("S2", "Legende / Notizen", bold)
setc("S3", "Hellblau = Eingabefeld  •  Schwarz = berechnet  •  Grün = Ergebnis", note_f)

# ============ LEFT BLOCK B:E ============
# 1. Ankauf
sectionbar(5, "B", "E", "1. Ankauf  (€/m² in Spalte E)")
setc("B6", "Kaufpreis", base); inp("D6", 180000); setc("E6", "=D6/D7", base, fmt=EUR2, align="right")
setc("B7", "Wohnfläche", base); inp("D7", 60, fmt='0 "m²"'); setc("E7", "m²", base, align="right")
setc("B8", "Notar + Grundbuch", base); inp("C8", 0.015, fmt=PCT); setc("D8", "=D6*C8", base, fmt=EUR); setc("E8", "=D8/D7", base, fmt=EUR2, align="right")
setc("B9", "Makler", base); inp("C9", 0.0357, fmt=PCT); setc("D9", "=D6*C9", base, fmt=EUR); setc("E9", "=D9/D7", base, fmt=EUR2, align="right")
setc("B10", "Grunderwerbsteuer", base); inp("C10", 0.065, fmt=PCT, ); setc("D10", "=D6*C10", base, fmt=EUR); setc("E10", "=D10/D7", base, fmt=EUR2, align="right")
ws["C10"].comment = Comment("Grunderwerbsteuer NRW = 6,5%", "FlipperImmo")
setc("B11", "Kaufnebenkosten", bold); setc("D11", "=SUM(D8:D10)", bold, fmt=EUR); setc("E11", "=D11/D7", base, fmt=EUR2, align="right")
setc("B12", "Gesamtinvestition", bold, fill=res_fill); setc("C12", "", fill=res_fill); setc("D12", "=D6+D11", bold, fmt=EUR, fill=res_fill); setc("E12", "=D12/D7", base, fmt=EUR2, align="right", fill=res_fill)
setc("B13", "Gebäudeanteil (für AfA)", base); inp("C13", 0.80, fmt=PCT); setc("D13", "=D6*C13", base, fmt=EUR); setc("E13", "Gebäudewert", note_f, align="right")
ws["C13"].comment = Comment("Anteil des Kaufpreises, der auf das Gebäude entfällt (nicht Grund & Boden). Nur dieser wird abgeschrieben. Typ. 75-85%.", "FlipperImmo")

# 2. Finanzierung
sectionbar(15, "B", "E", "2. Finanzierung")
setc("B16", "Finanzierungsquote", base); inp("C16", 1.00, fmt=PCT); setc("D16", "von Gesamtinvest.", note_f)
ws["C16"].comment = Comment("100% = Vollfinanzierung inkl. Nebenkosten (110%-Finanzierung). 90% = du bringst 10% Eigenkapital ein.", "FlipperImmo")
setc("B17", "Darlehenssumme", bold); setc("D17", "=D12*C16", bold, fmt=EUR)
setc("B18", "Eigenkapital", base); setc("D18", "=D12-D17", base, fmt=EUR)
setc("B19", "Sollzins p.a.", base); inp("C19", 0.040, fmt=PCT)
setc("B20", "anfängl. Tilgung p.a.", base); inp("C20", 0.020, fmt=PCT)
setc("B21", "Annuität p.a.", base); setc("D21", "=D17*(C19+C20)", base, fmt=EUR)
setc("B22", "Annuität p.M. (Kapitaldienst)", bold); setc("D22", "=D21/12", bold, fmt=EUR)
setc("B23", "  davon Zins p.M. (Jahr 1)", base); setc("D23", "=D17*C19/12", base, fmt=EUR)
setc("B24", "  davon Tilgung p.M.", base); setc("D24", "=D22-D23", base, fmt=EUR)

# 3. Mieteinnahmen
sectionbar(26, "B", "E", "3. Mieteinnahmen")
setc("B27", "Kaltmiete €/m²", base); inp("C27", 9.00, fmt=EUR2)
setc("B28", "Stellplatz / Sonstiges p.M.", base); inp("D28", 0, fmt=EUR)
setc("B29", "Kaltmiete p.M.", bold); setc("D29", "=C27*D7+D28", bold, fmt=EUR)
setc("B30", "Kaltmiete p.a.", bold); setc("D30", "=D29*12", bold, fmt=EUR)

# ============ RIGHT BLOCK G:I ============
# 4. Laufende Kosten
sectionbar(5, "G", "I", "4. Laufende Kosten (nicht umlagefähig, p.M.)")
setc("G6", "nicht uml. Hausgeld", base); inp("I6", 50, fmt=EUR)
ws["I6"].comment = Comment("Nur der NICHT umlagefähige Teil des Hausgelds (z.B. Verwaltung, Instandhaltungsrücklage der WEG). Umlagefähige Kosten trägt der Mieter.", "FlipperImmo")
setc("G7", "Instandhaltung €/m²/Jahr", base); inp("H7", 10, fmt=EUR2); setc("I7", "=H7*D7/12", base, fmt=EUR)
setc("G8", "Verwaltung (extern) p.M.", base); inp("I8", 25, fmt=EUR)
setc("G9", "Mietausfallwagnis", base); inp("H9", 0.03, fmt=PCT); setc("I9", "=D29*H9", base, fmt=EUR)
ws["H9"].comment = Comment("Puffer für Leerstand & Mietausfall, in % der Kaltmiete. Üblich 2-5%.", "FlipperImmo")
setc("G10", "Summe lfd. Kosten p.M.", bold); setc("I10", "=SUM(I6:I9)", bold, fmt=EUR)
setc("G11", "Summe lfd. Kosten p.a.", base); setc("I11", "=I10*12", base, fmt=EUR)

# 5. Cashflow
sectionbar(13, "G", "I", "5. Cashflow (Einnahmen − Ausgaben)")
setc("G14", "Kaltmiete p.M.", base); setc("I14", "=D29", base, fmt=EUR)
setc("G15", "− Kapitaldienst (Zins+Tilgung)", base); setc("I15", "=-D22", base, fmt=EUR)
setc("G16", "− lfd. Kosten", base); setc("I16", "=-I10", base, fmt=EUR)
setc("G17", "Cashflow VOR Steuer p.M.", white_bold, fill=sec_fill); setc("H17","",fill=sec_fill); setc("I17", "=I14+I15+I16", white_bold, fmt=EUR, fill=sec_fill, align="right")
setc("G18", "Cashflow vor Steuer p.a.", base); setc("I18", "=I17*12", base, fmt=EUR)
setc("G19", "AfA p.a. (2% Gebäude)", base); inp("H19", 0.02, fmt=PCT); setc("I19", "=D13*H19", base, fmt=EUR)
setc("G20", "Grenzsteuersatz", base); inp("H20", 0.42, fmt=PCT)
ws["H20"].comment = Comment("Dein persönlicher Grenzsteuersatz (inkl. Soli ca. 26-47%). Spitzensteuersatz 42%.", "FlipperImmo")
setc("G21", "steuerl. Ergebnis p.a.", base); setc("I21", "=D30-D23*12-I19-I11", base, fmt=EUR)
ws["I21"].comment = Comment("Kaltmiete − Zinsen − AfA − lfd. Kosten. Tilgung ist NICHT absetzbar. Negativ = Verlust mindert deine Steuer.", "FlipperImmo")
setc("G22", "Steuer p.a. (+Last / −Erstattung)", base); setc("I22", "=I21*H20", base, fmt=EUR)
setc("G23", "Cashflow NACH Steuer p.a.", white_bold, fill=sec_fill); setc("H23","",fill=sec_fill); setc("I23", "=I18-I22", white_bold, fmt=EUR, fill=sec_fill, align="right")
setc("G24", "Cashflow nach Steuer p.M.", bold, fill=res_fill); setc("H24","",fill=res_fill); setc("I24", "=I23/12", bold, fmt=EUR, fill=res_fill, align="right")

# 6. Renditen & Kennzahlen
sectionbar(26, "G", "I", "6. Renditen & Kennzahlen")
setc("G27", "Bruttomietrendite", base); setc("I27", "=D30/D6", bold, fmt=PCT2, align="right")
ws["I27"].comment = Comment("Jahreskaltmiete / Kaufpreis. Faustregel: ab ~4-5% interessant.", "FlipperImmo")
setc("G28", "Nettomietrendite", base); setc("I28", "=(D30-I11)/D12", bold, fmt=PCT2, align="right")
ws["I28"].comment = Comment("(Jahreskaltmiete − lfd. Kosten) / Gesamtinvestition inkl. Nebenkosten.", "FlipperImmo")
setc("G29", "Eigenkapitalrendite (n. St.)", base); setc("I29", '=IF(D18<=0,"n/a (Vollfin.)",(I23+D24*12)/D18)', bold, fmt=PCT2, align="right")
ws["I29"].comment = Comment("(Cashflow nach Steuer + Tilgung) / eingesetztes Eigenkapital. Bei 100%-Finanzierung nicht definiert.", "FlipperImmo")
setc("G30", "Kaufpreisfaktor", base); setc("I30", "=D6/D30", bold, fmt=NUM2, align="right")
ws["I30"].comment = Comment("Kaufpreis / Jahreskaltmiete. Kehrwert der Bruttorendite. Niedrig = günstig (z.B. 20x = 5%).", "FlipperImmo")
setc("G31", "Tilgung baut Vermögen p.a.", base); setc("I31", "=D24*12", base, fmt=EUR, align="right")

# Bewertung / Ampel
sectionbar(33, "G", "I", "7. Bewertung")
setc("G34", "Cashflow vor Steuer", base)
setc("I34", '=IF(I17>=50,"positiv ✓",IF(I17>=0,"knapp ±",​"negativ ✗"))'.replace("​",""), bold, align="right")
setc("G35", "Bruttomietrendite", base)
setc("I35", '=IF(I27>=0.045,"gut ✓",IF(I27>=0.035,"ok ±","niedrig ✗"))', bold, align="right")
setc("G36", "Gesamturteil", bold, fill=res_fill); setc("H36","",fill=res_fill)
setc("I36", '=IF(AND(I17>=0,I27>=0.04),"Kaufen prüfen ✓",IF(I17>=-100,"Grenzfall – verhandeln","Eher nein ✗"))', bold, align="right", fill=res_fill)

# ============ SZENARIO-/VERHANDLUNGSTABELLE ============
sectionbar(39, "B", "I", "8. Szenario: Kaufpreis-Verhandlung")
hdr = ["Discount", "Kaufpreis", "Gesamtinvest.", "Darlehen", "Annuität p.M.", "CF vor St. p.M.", "Bruttorendite", "Nettorendite"]
cols = ["B","C","D","E","F","G","H","I"]
for col, h in zip(cols, hdr):
    setc(f"{col}40", h, white_bold, fill=hdr_fill, align="center", bd=True)
    ws[f"{col}40"].font = bold
discounts = [0.0, 0.05, 0.10, 0.15, 0.20]
r = 41
for d in discounts:
    inp(f"B{r}", d, fmt=PCT)
    setc(f"C{r}", f"=$D$6*(1-B{r})", base, fmt=EUR, bd=True, align="right")
    setc(f"D{r}", f"=C{r}*(1+$C$8+$C$9+$C$10)", base, fmt=EUR, bd=True, align="right")
    setc(f"E{r}", f"=D{r}*$C$16", base, fmt=EUR, bd=True, align="right")
    setc(f"F{r}", f"=E{r}*($C$19+$C$20)/12", base, fmt=EUR, bd=True, align="right")
    setc(f"G{r}", f"=$D$29-F{r}-$I$10", base, fmt=EUR, bd=True, align="right")
    setc(f"H{r}", f"=$D$30/C{r}", base, fmt=PCT2, bd=True, align="right")
    setc(f"I{r}", f"=($D$30-$I$11)/D{r}", base, fmt=PCT2, bd=True, align="right")
    r += 1

# ---- Notizen / Anleitung (S-Spalte) ----
notes = [
 (5, "So nutzt du den Rechner:"),
 (6, "1. Trage nur die HELLBLAUEN Felder ein (Kaufpreis, Fläche, Zins, Miete ...)."),
 (7, "2. Alles andere rechnet sich automatisch."),
 (8, "3. Schau auf Block 5 (Cashflow) und 6 (Renditen) – das ist deine Kaufentscheidung."),
 (10, "Kapitaldienst = Zins + Tilgung. Das ist deine echte Monatsrate an die Bank."),
 (11, "Tilgung ist KEIN Verlust – du baust damit Eigenkapital/Vermögen auf (Block 6)."),
 (13, "Cashflow vor Steuer = Kaltmiete − Rate − lfd. Kosten."),
 (14, "Positiver Cashflow = die Immobilie trägt sich selbst."),
 (16, "Steuer: oft entsteht durch Zinsen + AfA ein steuerlicher Verlust → Erstattung,"),
 (17, "d.h. der Cashflow nach Steuer ist meist besser als vor Steuer."),
 (19, "Bruttomietrendite ab ~4-5% gilt als interessant (lagenabhängig)."),
 (20, "Kaufpreisfaktor < 25x ist solide, < 20x günstig."),
 (22, "Block 8: Wie ändern sich Cashflow & Rendite, wenn du den Preis drückst?"),
 (24, "Hinweis: vereinfachtes Modell, keine Steuer-/Anlageberatung."),
]
for row, txt in notes:
    setc(f"S{row}", txt, note_f)

# ---- column widths ----
widths = {"A":2,"B":24,"C":11,"D":13,"E":12,"F":4,"G":26,"H":11,"I":13,"J":3,"S":70}
from openpyxl.utils import get_column_letter
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.sheet_view.showGridLines = False
ws.freeze_panes = "A4"

# =====================================================================
# SHEET 2: PROJEKTION (Tilgungsplan + Zins-Sensitivität)
# =====================================================================
p = wb.create_sheet("Projektion")
from openpyxl.utils import column_index_from_string, get_column_letter

def psetc(coord, val, font=base, fill=None, fmt=None, align=None, bd=False, comment=None):
    c = p[coord]
    c.value = val
    c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if align: c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(align=="wrap"))
    if bd: c.border = border
    if comment: c.comment = Comment(comment, "FlipperImmo")
    return c

def pinp(coord, val, fmt):
    return psetc(coord, val, font=Font(name=FONT, size=10, color="00008B"), fill=input_fill, fmt=fmt, align="right", bd=True)

def psectionbar(row, col_start, col_end, text):
    cs = p[f"{col_start}{row}"]; cs.value = text; cs.font = sec
    a = column_index_from_string(col_start); b = column_index_from_string(col_end)
    for i in range(a, b+1):
        p[f"{get_column_letter(i)}{row}"].fill = sec_fill
    p.merge_cells(f"{col_start}{row}:{col_end}{row}")

from openpyxl.utils import column_index_from_string
psetc("B2", "20-Jahres-Projektion: Tilgungsplan, Cashflow & Zins-Sensitivität", title)
psetc("B3", "Annuitätendarlehen. Werte ziehen aus Blatt 'Kalkulation'. Hellblau = anpassbar.", note_f)

# Annahmen (Eingaben)
psectionbar(5, "B", "C", "Annahmen")
psetc("B6", "Mietsteigerung p.a.", base); pinp("C6", 0.015, PCT)
psetc("B7", "Kostensteigerung p.a.", base); pinp("C7", 0.020, PCT)
psetc("B8", "Zinsbindung (Jahre)", base); pinp("C8", 10, '0')
psetc("B9", "Anschlusszins p.a.", base); pinp("C9", 0.050, PCT)
p["C9"].comment = Comment("Zins-Sensitivität: ändere diesen Wert (z.B. 5%, 6%, 7%) und schau, wie Restschuld & Cashflow ab Jahr 11 reagieren.", "FlipperImmo")

# Herleitung (aus Kalkulation)
psectionbar(5, "E", "G", "Herleitung (aus Kalkulation)")
psetc("E6", "Darlehen", base); psetc("G6", "=Kalkulation!D17", base, fmt=EUR, align="right")
psetc("E7", "Sollzins (Phase 1)", base); psetc("G7", "=Kalkulation!C19", base, fmt=PCT, align="right")
psetc("E8", "anf. Tilgung", base); psetc("G8", "=Kalkulation!C20", base, fmt=PCT, align="right")
psetc("E9", "Annuität Phase 1 p.a.", base); psetc("G9", "=G6*(G7+G8)", base, fmt=EUR, align="right")
psetc("E10", "Restschuld bei Anschluss", base); psetc("G10", "=G6*(1+G7)^C8-G9*((1+G7)^C8-1)/G7", base, fmt=EUR, align="right")
psetc("E11", "Annuität Phase 2 p.a.", base); psetc("G11", "=G10*(C9+G8)", base, fmt=EUR, align="right")

# Tilgungsplan-Tabelle
psectionbar(13, "A", "L", "Tilgungsplan & Cashflow (20 Jahre)")
hdr2 = ["Jahr","Restschuld\nAnfang","Zinssatz","Zins","Tilgung","Restschuld\nEnde","Annuität","Kaltmiete","lfd. Kosten","CF vor St.","CF n. St.","CF kumul.\n(n. St.)"]
cols2 = ["A","B","C","D","E","F","G","H","I","J","K","L"]
for col, h in zip(cols2, hdr2):
    psetc(f"{col}14", h, bold, fill=hdr_fill, align="center", bd=True)
    p[f"{col}14"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

start = 15
for i in range(20):
    r = start + i
    yr = i + 1
    psetc(f"A{r}", yr if i == 0 else f"=A{r-1}+1", base, fmt='0', align="center", bd=True)
    psetc(f"B{r}", "=$G$6" if i == 0 else f"=F{r-1}", base, fmt=EUR, bd=True, align="right")
    psetc(f"C{r}", f"=IF(A{r}<=$C$8,$G$7,$C$9)", base, fmt=PCT, bd=True, align="right")
    psetc(f"D{r}", f"=B{r}*C{r}", base, fmt=EUR, bd=True, align="right")
    psetc(f"G{r}", f"=MIN(IF(A{r}<=$C$8,$G$9,$G$11),B{r}+D{r})", base, fmt=EUR, bd=True, align="right")
    psetc(f"E{r}", f"=G{r}-D{r}", base, fmt=EUR, bd=True, align="right")
    psetc(f"F{r}", f"=MAX(0,B{r}-E{r})", base, fmt=EUR, bd=True, align="right")
    psetc(f"H{r}", f"=Kalkulation!$D$30*(1+$C$6)^(A{r}-1)", base, fmt=EUR, bd=True, align="right")
    psetc(f"I{r}", f"=Kalkulation!$I$11*(1+$C$7)^(A{r}-1)", base, fmt=EUR, bd=True, align="right")
    psetc(f"J{r}", f"=H{r}-G{r}-I{r}", base, fmt=EUR, bd=True, align="right")
    # CF nach Steuer: CF vor St − Steuer(steuerl. Ergebnis = Kaltmiete − Zins − AfA − lfd Kosten)
    psetc(f"K{r}", f"=J{r}-((H{r}-D{r}-Kalkulation!$I$19-I{r})*Kalkulation!$H$20)", base, fmt=EUR, bd=True, align="right")
    psetc(f"L{r}", f"=K{r}" if i == 0 else f"=L{r-1}+K{r}", base, fmt=EUR, bd=True, align="right")

# Zusammenfassung
endrow = start + 19
psectionbar(37, "B", "G", "Auf einen Blick")
psetc("B38", "Restschuld nach 10 Jahren", base); psetc("D38", f"=F{start+9}", bold, fmt=EUR, align="right")
psetc("B39", "Restschuld nach 20 Jahren", base); psetc("D39", f"=F{endrow}", bold, fmt=EUR, align="right")
psetc("B40", "getilgt nach 20 J.", base); psetc("D40", f"=$G$6-F{endrow}", base, fmt=EUR, align="right")
psetc("B41", "Cashflow kumuliert (20 J., n. St.)", base); psetc("D41", f"=L{endrow}", bold, fmt=EUR, align="right")
psetc("B42", "Vermögenszuwachs 20 J. (Tilgung + CF kum.)", bold, fill=res_fill); psetc("C42","",fill=res_fill); psetc("D42", f"=($G$6-F{endrow})+L{endrow}", bold, fmt=EUR, align="right", fill=res_fill)
p["D42"].comment = Comment("Ohne Wertsteigerung der Immobilie. Reiner Effekt aus Tilgung (Vermögensaufbau) plus kumuliertem Cashflow nach Steuer.", "FlipperImmo")

pnotes = [
 (38, "→ Zins-Sensitivität: oben den 'Anschlusszins' (C9) ändern."),
 (39, "   Ab Jahr 11 (nach Zinsbindung) rechnet die Tabelle mit dem neuen Zins."),
 (40, "   So siehst du, was bei steigenden Zinsen mit deinem Cashflow passiert."),
 (42, "Hinweis: ohne Immobilien-Wertsteigerung gerechnet (konservativ)."),
]
for row, txt in pnotes:
    psetc(f"I{row}", txt, note_f)

pwidths = {"A":6,"B":13,"C":10,"D":11,"E":10,"F":13,"G":11,"H":11,"I":11,"J":11,"K":11,"L":12}
for col, w in pwidths.items():
    p.column_dimensions[col].width = w
p.sheet_view.showGridLines = False
p.freeze_panes = "A15"

out = r"C:\Users\HP\Documents\Claude\immobilien-rechner\VermietRechner.xlsx"
import os
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("saved", out)
