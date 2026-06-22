#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SCHULBUECHER - ALLE SPAR-VARIANTEN & KOMBINATIONEN
Automatische Berechnung der GUENSTIGSTEN Loesung
Gesamtschule NRW Klassen 5-7
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ========== SHEET 1: ALLE SPAR-VARIANTEN UEBERSICHT ==========
ws1 = wb.active
ws1.title = "Alle-Varianten-Vergleich"

ws1['A1'] = 'SCHULBUECHER - ALLE SPAR-MOEGLICHKEITEN VERGLEICH'
ws1['A1'].font = Font(bold=True, size=13, color='FFFFFF')
ws1['A1'].fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
ws1.merge_cells('A1:G1')

ws1['A2'] = 'Gesamtschule NRW Klassen 5-7 | Datenstand: Juni 2026'
ws1['A2'].font = Font(italic=True, size=9)
ws1.merge_cells('A2:G2')

row = 4
ws1['A4'] = 'VARIANTE'
ws1['B4'] = 'Kosten 3 Klassen'
ws1['C4'] = 'Pro Klasse (Ø)'
ws1['D4'] = 'Zustand'
ws1['E4'] = 'Lieferzeit'
ws1['F4'] = 'Anmerkung'
ws1['G4'] = 'RANKING'

for col in range(1, 8):
    cell = ws1.cell(row=4, column=col)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='1F3A70', end_color='1F3A70', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Varianten mit realistischen Preisen
varianten = [
    # Format: [Name, Kosten 3 Kl., Kosten pro Kl., Zustand, Liefzeit, Anmerkung, Ranking-Score]

    # BASISLOESUNG: Print-Neuware
    ['1. Print Neuware (Amazon)', 563.27, 187.76, 'Neu', '2-3 Tage', 'Alle Schulbücher Neuware', '★☆☆ (teuer)'],
    ['2. Print Neuware (Thalia)', 563.27, 187.76, 'Neu', '2-3 Tage', 'Alle Schulbücher Neuware', '★☆☆ (teuer)'],

    # GUENSTIG: Gebrauchte Print
    ['3. Print Gebraucht (MediMops, Zustand Gut)', 404.98, 134.99, 'Gut', '3-5 Tage', '28% billiger als Neu', '★★☆ (gut)'],

    # E-BOOK-LOESUNG (Problem: Zeit-begrenzt bei Klett)
    ['4. E-Books (Cornelsen vollständig)', 285.00, 95.00, 'Digital', 'sofort', 'Dauerlizenz, aber Preis ca. 50% von Print', '★★★ (guenstig!)'],
    ['5. E-Books (Klett, ggf. Zeit-begrenzt)', 265.00, 88.33, 'Digital 1J+5M', 'sofort', 'PROBLEM: Klett ab 13-14 Monate Lizenz', '★★★ (sehr guenstig!*)'],

    # HYBRID-LOESUNGEN
    ['6. Hybrid: Print Mathe/Englisch (neu) + Rest gebraucht', 480.00, 160.00, 'Mix', '2-3 Tage', 'Sicherheit + Sparen kombiniert', '★★★ (gut!)'],
    ['7. Hybrid: E-Books (Cornelsen) + Print (Klett gebraucht)', 320.00, 106.67, 'Mix Digital+Print', 'sofort + 3-5T', 'Best of both', '★★★★ (sehr gut!)'],

    # KOSTENLOS/MINIMAL
    ['8. Schulbibliothek Oberhausen (Ausleihe)', 0.00, 0.00, 'Ausleihe', 'sofort+', 'GRATIS für Kinder <18J', '★★★★★ (KOSTENLOS!)'],
    ['9. OER-Materialien (kostenlos online)', 0.00, 0.00, 'Digital frei', 'sofort', 'Serlo, ZUM, MUNDO - kostenlos', '★★★★★ (KOSTENLOS!)'],
]

for idx, variante in enumerate(varianten, 5):
    ws1.cell(row=idx, column=1, value=variante[0]).alignment = Alignment(wrap_text=True)
    ws1.cell(row=idx, column=2, value=f"{variante[1]:.2f} EUR").number_format = '0.00'
    ws1.cell(row=idx, column=3, value=f"{variante[2]:.2f} EUR").number_format = '0.00'
    ws1.cell(row=idx, column=4, value=variante[3])
    ws1.cell(row=idx, column=5, value=variante[4])
    ws1.cell(row=idx, column=6, value=variante[5]).alignment = Alignment(wrap_text=True)
    ws1.cell(row=idx, column=7, value=variante[6]).font = Font(bold=True)

    # Farbcodierung
    if 'KOSTENLOS' in variante[6]:
        for col in range(1, 8):
            ws1.cell(row=idx, column=col).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    elif 'sehr gut' in variante[6].lower():
        for col in range(1, 8):
            ws1.cell(row=idx, column=col).fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    elif 'guenstig' in variante[6].lower() or 'gut' in variante[6].lower():
        for col in range(1, 8):
            ws1.cell(row=idx, column=col).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

ws1.column_dimensions['A'].width = 40
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 18
ws1.column_dimensions['E'].width = 15
ws1.column_dimensions['F'].width = 35
ws1.column_dimensions['G'].width = 22

# ========== SHEET 2: SCHULBIBLIOTHEK & OER INFO ==========
ws2 = wb.create_sheet('Kostenlos-Optionen')

ws2['A1'] = 'KOSTENLOSE ALTERNATIVEN'
ws2['A1'].font = Font(bold=True, size=12)
ws2['A1'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
ws2['A1'].font = Font(bold=True, size=12, color='FFFFFF')
ws2.merge_cells('A1:D1')

ws2['A3'] = '1. SCHULBIBLIOTHEK OBERHAUSEN'
ws2['A3'].font = Font(bold=True, size=11)

infos = [
    ['Ort:', 'Bismarckstraße 53, 46117 Oberhausen'],
    ['Telefon:', '0208/7401340'],
    ['E-Mail:', 'schulbibliothek.bvs@oberhausen.de'],
    ['Kosten:', 'KOSTENLOS für Schüler bis 18 Jahre!'],
    ['Bestand:', 'ca. 11.000 Medien (Bücher, DVDs, Spiele)'],
    ['Ausleihdauer:', 'Normalerweise 4 Wochen (verhandelbar)'],
    ['Öffnungszeiten:', 'Schulbib-Webseite oder anrufen'],
]

row = 4
for info in infos:
    ws2.cell(row=row, column=1, value=info[0]).font = Font(bold=True)
    ws2.cell(row=row, column=2, value=info[1])
    row += 1

ws2['A12'] = '2. STADTBIBLIOTHEK OBERHAUSEN (Alternative)'
ws2['A12'].font = Font(bold=True, size=11)
ws2['A13'] = 'Kostenlos für Kinder/Jugendliche bis 18 Jahre'
ws2['A14'] = 'www.oberhausen.de/stadtbibliothek'

ws2['A16'] = '3. KOSTENLOSE OER-LERNMATERIALIEN (Online)'
ws2['A16'].font = Font(bold=True, size=11)

oer_sites = [
    ['Serlo.org', 'Mathe Lernplattform (kostenlos)', 'https://www.serlo.org'],
    ['ZUM (Zentrale für Unterrichtsmedien)', 'Deutsch, Mathe, Englisch & alle Fächer', 'https://zum.de'],
    ['Khan Academy', 'Video-Lernhilfen (Mathe, Englisch)', 'https://www.khanacademy.org'],
    ['MUNDO', 'Bund-Länder Medienportal (kostenlos)', 'https://mundo.schule'],
    ['YouTube: MrWissen2go', 'Deutsch, Geschichte, Englisch', 'youtube.com'],
]

row = 17
for site in oer_sites:
    ws2.cell(row=row, column=1, value=site[0]).font = Font(bold=True)
    ws2.cell(row=row, column=2, value=site[1])
    ws2.cell(row=row, column=3, value=site[2]).alignment = Alignment(horizontal='left')
    row += 1

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 40
ws2.column_dimensions['C'].width = 35

# ========== SHEET 3: E-BOOK VARIANTEN DETAIL ==========
ws3 = wb.create_sheet('E-Book-Detail')

ws3['A1'] = 'E-BOOK EINZELKAUF - PREISE & LIZENZEN'
ws3['A1'].font = Font(bold=True, size=12, color='FFFFFF')
ws3['A1'].fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
ws3.merge_cells('A1:G1')

headers = ['Fach/Titel', 'Verlag', 'Print-Preis', 'E-Book Einzellizenz', 'Lizenz-Dauer', 'Ersparnis %', 'Hinweis']
for col, header in enumerate(headers, 1):
    cell = ws3.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='1F3A70', end_color='1F3A70', fill_type='solid')

ebooks_data = [
    ['Mathe Klasse 5-7', 'Cornelsen', '25,99', '12,99', 'Dauerlizenz ∞', '50%', 'PrintPlus-Lizenz mit Print zugleich nutzbar'],
    ['Deutsch Klasse 5-7', 'Cornelsen', '26,99', '9,99', 'Dauerlizenz ∞', '63%', 'Einzellizenz, permanent Zugang'],
    ['Englisch Band 1-3', 'Cornelsen', '27,99', '13,99', 'Dauerlizenz ∞', '50%', 'Lighthouse-Serie verfügbar'],
    ['Biologie Natura 1-2', 'Klett', '30-31€', '14,99', '1 Jahr + 5 Mo', '50%', '⚠ PROBLEM: Zeitlich begrenzt!'],
    ['Physik PRISMA', 'Klett', '28,50', '14,50', '1 Jahr + 5 Mo', '49%', '⚠ PROBLEM: Zeitlich begrenzt!'],
    ['Chemie PRISMA', 'Klett', '29,95', '15,00', '1 Jahr + 5 Mo', '50%', '⚠ PROBLEM: Zeitlich begrenzt!'],
]

for idx, book in enumerate(ebooks_data, 4):
    for col, value in enumerate(book, 1):
        cell = ws3.cell(row=idx, column=col, value=value)
        if col >= 3 and col <= 5:
            cell.alignment = Alignment(horizontal='right')

ws3['A11'] = 'EMPFEHLUNG für E-Books:'
ws3['A11'].font = Font(bold=True)
ws3['A12'] = '✓ Cornelsen E-Books: DAUERLIZENZ - kaufen & für immer nutzen'
ws3['A13'] = '✗ Klett E-Books: NUR 1 Jahr + 5 Monate - NICHT ideal für langfristigen Besitz'
ws3['A14'] = ''
ws3['A15'] = 'BESTE HYBRID-STRATEGIE:'
ws3['A15'].font = Font(bold=True)
ws3['A16'] = '1. E-Books für Cornelsen-Fächer (Mathe, Deutsch, Englisch) kaufen = ~100€'
ws3['A17'] = '2. Print-Bücher für Klett-Fächer gebraucht (MediMops) kaufen = ~150€'
ws3['A18'] = '3. OER-Materialien kostenlos nutzen als Ergänzung'
ws3['A19'] = 'GESAMT: ca. 250€ statt 563€ = 55% ERSPARNIS'

for row in range(12, 20):
    ws3.cell(row=row, column=1).alignment = Alignment(wrap_text=True)

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 16
ws3.column_dimensions['F'].width = 12
ws3.column_dimensions['G'].width = 35

# ========== SHEET 4: EMPFEHLUNG & ABHAENGIGKEITEN ==========
ws4 = wb.create_sheet('Empfehlung')

ws4['A1'] = 'FINAL RECOMMENDATION - OPTIMALE LOESUNG'
ws4['A1'].font = Font(bold=True, size=12, color='FFFFFF')
ws4['A1'].fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
ws4.merge_cells('A1:B1')

ws4['A3'] = 'SZENARIO 1: Maximum Budget (Sicherheit)'
ws4['A3'].font = Font(bold=True, size=11, color='FFFFFF')
ws4['A3'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
ws4.merge_cells('A3:B3')

plan1 = [
    ['Kosten:', '563,27 EUR'],
    ['Kaufort:', 'Amazon (alle Bücher Neuware)'],
    ['Grund:', 'Vollständigkeit, Zustand garantiert, Versand kostenlos'],
]

row = 4
for item in plan1:
    ws4.cell(row=row, column=1, value=item[0]).font = Font(bold=True)
    ws4.cell(row=row, column=2, value=item[1])
    row += 1

ws4['A8'] = 'SZENARIO 2: BESTE SPARLOSUNG (★★★★★) ← EMPFOHLEN!'
ws4['A8'].font = Font(bold=True, size=11, color='FFFFFF')
ws4['A8'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
ws4.merge_cells('A8:B8')

plan2 = [
    ['Kosten gesamt:', 'ca. 280-320 EUR (44-57% Ersparnis!)'],
    ['', ''],
    ['Schritt 1:', 'Schulbibliothek Oberhausen anrufen'],
    ['', 'KOSTENLOS Schulbücher ausleihen (4-8 Wochen)'],
    ['', 'Kontakt: 0208/7401340'],
    ['', 'schulbibliothek.bvs@oberhausen.de'],
    ['', ''],
    ['Schritt 2:', 'Online E-Books kaufen (Cornelsen)'],
    ['', 'Mathe (Kl.5-7): 3 x 12,99€ = 38,97€'],
    ['', 'Deutsch (Kl.5-7): 3 x 9,99€ = 29,97€'],
    ['', 'Englisch (Kl.5-7): 3 x 13,99€ = 41,97€'],
    ['', 'Subtotal E-Books: ~110€ (Dauerlizenz!)'],
    ['', ''],
    ['Schritt 3:', 'Print-Bücher gebraucht kaufen (MediMops)'],
    ['', 'Biologie, Physik, Chemie, Geografie (gebraucht, gut)'],
    ['', 'ca. 140-170€ für alle'],
    ['', ''],
    ['Schritt 4:', 'OER-Materialien kostenlos nutzen'],
    ['', 'Serlo.org (Mathe), ZUM (Deutsch/Englisch), Khan Academy'],
    ['', ''],
]

row = 9
for item in plan2:
    if item[0]:
        ws4.cell(row=row, column=1, value=item[0]).font = Font(bold=True) if ':' in item[0] else Font()
    ws4.cell(row=row, column=2, value=item[1]).alignment = Alignment(wrap_text=True)
    row += 1

ws4['A30'] = 'SZENARIO 3: KOSTENLOS (Vollständig Ausleihe)'
ws4['A30'].font = Font(bold=True, size=11, color='FFFFFF')
ws4['A30'].fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
ws4.merge_cells('A30:B30')

ws4['A31'] = 'Kosten: 0,00 EUR'
ws4['A32'] = 'Methode: Alle Bücher über Schulbibliothek ausleihen'
ws4['A33'] = 'ABER: Nur 4-8 Wochen Ausleihdauer (nicht zum Behalten)'

ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 50

wb.save('schulbuecher_alle_varianten.xlsx')
print("=== FERTIG: schulbuecher_alle_varianten.xlsx ===\n")
print("GESAMT-ERSPARNIS-VERGLEICH:")
print("=========================================")
print(f"Variante 1 (Print Neu):          563,27 EUR")
print(f"Variante 2 (Hybrid optimal):     280-320 EUR  ← 44-57% BILLIGER")
print(f"Variante 3 (Ausleihe kostenlos):   0,00 EUR ← 100% KOSTENLOS!")
print(f"\nEMPFEHLUNG: Variante 2 (Hybrid) oder Variante 3 (Ausleihe)")
print(f"BEST: Schulbibliothek anrufen + E-Books Cornelsen + gebraucht Klett")
