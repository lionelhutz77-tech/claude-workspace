#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Preise in Excel eintragen - Recherche-Ergebnisse
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Datei laden
wb = load_workbook('schulbuecher_preisvergleich.xlsx')
ws3 = wb['Kalkulation']

# Gefundene Preise eintragen (aus WebSearch-Ergebnissen)
preise = [
    # Format: [Titel, ISBN, Menge, Amazon (neu), MediMops (gebraucht), Thalia, Günstigster]
    # Klasse 5
    ['Dreifach Mathe 5 (Schulbuch)', '978-3-06-043746-7', 1, 26.99, 'recherchieren', 'recherchieren'],
    ['Dreifach Mathe 5 (Arbeitsheft)', '978-3-06-043724-5', 1, 9.99, 'recherchieren', 'recherchieren'],
    ['Deutschbuch 5', '978-3-06-062085-0', 1, 'recherchieren', 15.00, 'recherchieren'],
    ['Lighthouse 1 (Band 1)', '978-3-06-032373-9', 1, 27.99, 'recherchieren', 27.99],
    ['Natura 1 (Bio Kl.5/6)', '978-3-12-049200-9', 1, 30.50, 'recherchieren', 30.50],
    ['Geografie 5', '978-3-06-040572-5', 1, 'recherchieren', 'recherchieren', 'recherchieren'],
    ['PRISMA Physik 5/6', '978-3-12-068781-8', 1, 'recherchieren', 'recherchieren', 'recherchieren'],

    # Klasse 6
    ['Dreifach Mathe 6 (Schulbuch)', '978-3-06-043747-4', 1, 26.99, 'recherchieren', 'recherchieren'],
    ['Dreifach Mathe 6 (Arbeitsheft)', '978-3-06-043725-2', 1, 9.99, 'recherchieren', 'recherchieren'],
    ['Deutschbuch 6', '978-3-06-063401-7', 1, 'recherchieren', 15.00, 'recherchieren'],
    ['Lighthouse 2 (Band 2)', '978-3-06-032409-5', 1, 27.99, 'recherchieren', 27.99],
    ['Natura 1 (noch)', '978-3-12-049200-9', 1, 30.50, 'recherchieren', 30.50],
    ['Geografie 6', '978-3-06-040572-5', 1, 'recherchieren', 'recherchieren', 'recherchieren'],
    ['PRISMA Physik 5/6', '978-3-12-068781-8', 1, 'recherchieren', 'recherchieren', 'recherchieren'],

    # Klasse 7
    ['Dreifach Mathe 7 (Schulbuch)', '978-3-06-043748-1', 1, 26.99, 'recherchieren', 'recherchieren'],
    ['Dreifach Mathe 7 (Arbeitsheft)', '978-3-06-043726-9', 1, 9.99, 'recherchieren', 'recherchieren'],
    ['Deutschbuch 7', '978-3-06-063402-4', 1, 'recherchieren', 15.00, 'recherchieren'],
    ['Lighthouse 3 (Band 3)', '978-3-06-032410-1', 1, 27.99, 'recherchieren', 27.99],
    ['Natura 2 (Bio 7-9)', '978-3-12-049207-8', 1, 31.50, 'recherchieren', 31.50],
    ['Geografie 7', '978-3-06-040572-5', 1, 'recherchieren', 'recherchieren', 'recherchieren'],
    ['PRISMA Chemie 7-10', '978-3-12-068509-8', 1, 'recherchieren', 'recherchieren', 'recherchieren'],
]

# Daten in Kalkulation-Sheet eintragen (ab Zeile 4)
for idx, preis_row in enumerate(preise, 4):
    for col_idx, value in enumerate(preis_row, 1):
        ws3.cell(row=idx, column=col_idx, value=value)

# Hinweis in gelb für "recherchieren"
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
for row in ws3.iter_rows(min_row=4, max_row=4+len(preise)-1, min_col=1, max_col=7):
    for cell in row:
        if cell.value == 'recherchieren':
            cell.fill = yellow_fill

wb.save('schulbuecher_preisvergleich.xlsx')
print("Preise eingetragen - warte auf WebSearch-Reset für weitere Recherche")
